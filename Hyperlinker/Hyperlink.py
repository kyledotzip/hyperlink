from openpyxl import Workbook, load_workbook
import random

def GetValues(file):
    wb = load_workbook('{}'.format(file))
    ws = wb.active
    ws = wb.worksheets[0]

    values = []

    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
        for cell in row:
            if cell.value is None:
                values.append(0)
            else:
                values.append(int(cell.value))
    return values

def Assign(Dictionary, InputArray, file):
    wb = load_workbook('{}'.format(file))
    ws = wb.active
    ws = wb.worksheets[0]

    for i, value in enumerate(InputArray):
        if value == 0:
            continue
        cell = ws.cell(row = i+2, column = 1)
        cell.hyperlink = Dictionary[value]
        cell.style = "Hyperlink"

    name = "RENAME"
    wb.save("{}.xlsx".format(name))
